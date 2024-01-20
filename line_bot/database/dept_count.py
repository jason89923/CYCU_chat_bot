import openpyxl
import matplotlib.pyplot as plt
from pprint import pprint
import matplotlib.font_manager as fm
import pymysql
import global_variable as gv

plt.rcParams['font.sans-serif'] = ['SimHei']  # 将字体设置为黑体

def count_dp():
    dept_list = { 
        '應用數學系' : '應數系',
        '物理學系' : '物理系',
        '化學系' : '化學系',
        '心理學系' : '心理系',
        '生物科技學系' : '生科系',
        '奈米學位學程' : '奈米學位學程',
        '工業與系統工程學系' : '工業系',
        '電子工程學系' : '電子系',
        '資訊工程學系' : '資訊系',
        '電機工程學系' : '電機系',
        '電機資訊學院人工智慧應用學士學位學程' : '人工智慧學士學程',
        '電機資訊學院智慧運算與大數據學士班' : '智慧運算大數據學士班',
        '電機資訊學院學士班' : '電資學院學士班',
        '中原威大工程雙學士' : '中原威大工程雙學士',
        '智慧運算大數據碩士' : '智慧運算大數據碩士',
        '化學工程學系' : '化工系',
        '土木工程學系' : '土木系',
        '機械工程學系' : '機械系',
        '生物醫學工程學系' : '醫工系',
        '環境工程學系' : '環工系',
        '企業管理學系' : '企管系',
        '國際經營與貿易學系' : '國貿系',
        '會計學系' : '會計系',
        '資訊管理學系' : '資管系',
        '財務金融學系' : '財金系',
        '國際商學學士學程' : '國際商學學士學程',
        '中原天普商學雙學士' : '中原天普商學雙學士',
        '巨量碩士學程' : '巨量碩士學程',
        '國際商碩' : '國際商碩',
        '商學博' : '商學博',
        '建築學系' : '建築系',
        '商業設計學系' : '商設系',
        '室內設計學系' : '室設系',
        '地景建築學系' : '地景建築學系',
        '設計學院設計學士原住民專班' : '設計學士原住民專班',
        '文創設計碩' : '文創設計碩',
        '設計博' : '設計博',
        '特殊教育學系' : '特教系',
        '應用外國語文學系' : '應外系',
        '應用華語文學系' : '應華系',
        '師培中心' : '師培中心',
        '人文與教育學院學士學位學程' : '人育學士學程',
        '外籍不分系' : '外籍不分系',
        '音樂產業碩士' : '音樂產業碩士',
        '宗研所' : '宗研所',
        '教研所' : '教研所',
        '財經法律學系' : '財法系'
    }
 
    connect_db = pymysql.connect(host=gv.SQL.host, port=gv.SQL.port, user=gv.SQL.user, passwd=gv.SQL.passwd, charset=gv.SQL.charset, db=gv.SQL.db)    
    cursor = connect_db.cursor()
    
    state = {
        '主選單':'主選單',
        '查詢課程':'指定',
        '丟水球':'丟水球',
        '系統改善':'系統改善'
    }
    
    client = gv.mongo_client
    db = client.line_bot
    std_info = db['std_course_info']
    dept_count = {}  # 创建一个空字典来存储部门记录数
    for dept in dept_list.keys():
        count = std_info.count_documents({'dep_name':{ '$regex' : dept}})  # 统计特定部门的记录数
        student_ids = std_info.distinct('stmd3_id', {'dep_name': { '$regex' : dept}})
        msg_count = 0
        function_count_list = [0 for _ in range(len(state))]
        for sid in student_ids:
            query = "SELECT COUNT(*) FROM system_log WHERE sid = %s"
            cursor.execute(query, sid)
            msg_count += cursor.fetchone()[0]
            for i,s in enumerate(state.values()):
                query = f"SELECT COUNT(*) FROM system_log WHERE state LIKE '%{s}%' and sid = {sid}"
                cursor.execute(query)
                function_count_list[i] += cursor.fetchone()[0]
        
        dept_count[dept] = [count,msg_count]  # 将部门记录数存储在字典中
        dept_count[dept].extend(function_count_list)
        print(dept,dept_count[dept])

    
    
    # 创建一个新的Excel工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # 将部门记录数写入Excel表格
    sheet.cell(row=1, column=1, value='Department')
    sheet.cell(row=1, column=2, value='Count')
    sheet.cell(row=1, column=3, value='MSG_Count')
    sheet.cell(row=1, column=4, value='主選單')
    sheet.cell(row=1, column=5, value='查詢課程')
    sheet.cell(row=1, column=6, value='丟水球')
    sheet.cell(row=1, column=7, value='系統改善')
    row_num = 2
    for dept, count in dept_count.items():
        sheet.cell(row=row_num, column=1, value=dept)
        sheet.cell(row=row_num, column=2, value=count[0])
        sheet.cell(row=row_num, column=3, value=count[1])
        sheet.cell(row=row_num, column=4, value=count[2])
        sheet.cell(row=row_num, column=5, value=count[3])
        sheet.cell(row=row_num, column=6, value=count[4])
        sheet.cell(row=row_num, column=7, value=count[5])
        row_num += 1

    # 保存Excel文件
    workbook.save('department_count.xlsx')

    
if __name__ == '__main__':
    count_dp()
        
        
